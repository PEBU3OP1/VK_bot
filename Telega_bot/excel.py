from datetime import datetime
import os.path
import time

import openpyxl


my_path = "/Users/sevak/PycharmProjects/Telega_bot/New.xlsx"
def money_saver(price, for_what, recipient_name):

    if os.path.isfile(my_path):
        wb = openpyxl.load_workbook(my_path)
        wb.active = 0
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        wb.active = 0
        sheet = wb.active
        sheet['A1'] = 'Цена'
        sheet['B1'] = 'За что'
        sheet['C1'] = 'Кому'





    a = 2

    while True:
        if sheet['A' + str(a)].value == None:
            sheet['A' + str(a)].value = price
            sheet['B' + str(a)].value = for_what
            sheet['C' + str(a)].value = recipient_name
            break
        else:
            a += 1





    wb.save(my_path)
    wb.close()

# while True:
#     now = datetime.now()
#     current_time = now.strftime('%H:%M:%S')
#     if current_time == '19:45:00':
#         print('Время пришло')
#         time.sleep(1)

def money_result_per_day():
    wb = openpyxl.load_workbook(my_path)
    wb.active = 0
    sheet = wb.active

    result = ''
    summary = 0
    name_list = []
    a = 2

    while sheet['C' + str(a)].value!= None:
        name_list.append(sheet['C' + str(a)].value)
        a += 1
    name_list = list(set(name_list))

    for i in name_list:
        a = 2
        while sheet['A' + str(a)].value!= None:
            if sheet['C' + str(a)].value == i:
                summary = summary + int(sheet['A' + str(a)].value)
            a += 1
        result = result + f'{i} {summary} \n'
        summary = 0
    return result
